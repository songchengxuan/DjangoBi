import datetime

import psycopg2
import psycopg2.extras
from django.forms import model_to_dict
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from app01 import models
from openpyxl import load_workbook
from DjangoBi.gplink import gp_connect

def index(request):
    return HttpResponse("欢迎使用")

# ======Y1.主推商品属性表=======
# 按品类查询
@require_http_methods(["GET"])
def y1_list(request):
    """ list列表 + 查找数据(模糊查询) """
    response = {}
    try:
        # === 查询开始 ===
        data_dict = {}
        main_attribute = request.GET.get('main_attribute')
        if main_attribute:
            data_dict["main_attribute__contains"] = main_attribute
        # === 查询结束 ===
        queryset = models.MainAttribute.objects.filter(**data_dict)
        response['data'] = []
        for i in queryset:
            response['data'].append(model_to_dict(i))
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y1_add(request):
    """ list列表添加与编辑 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        data_dict = {
            "branch_id": request.GET.get('branch_id'),
            "satnr": request.GET.get('satnr'),
            "main_attribute": request.GET.get('main_attribute'),
            "modify_time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        if nid:
            models.MainAttribute.objects.filter(id=nid).update(**data_dict)
        else:
            models.MainAttribute.objects.create(**data_dict)
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y1_delete(request):
    """ list列表删除 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        models.MainAttribute.objects.filter(id=nid).delete()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["POST"])
def y1_upload(request):
    """ Excel上传+表头认证 """
    response = {}
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.打开excel
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.认证表头
    colnames = []
    colnames_real = []
    for col in sheet.iter_cols(max_col=3):
        colnames.append(col[0].value)
    exclude_fields = ('id', 'create_time', 'modify_time')
    params = [f for f in models.MainAttribute._meta.fields if f.name not in exclude_fields]
    for para in params:
        colnames_real.append(para.verbose_name)
    if colnames != colnames_real:
        response['msg'] = 'Excel字段名称不匹配'
        return JsonResponse(response)
    # 4.写入数据库
    for row in sheet.iter_rows(min_row=2):
        data_dict = {
            "branch_id": row[0].value,
            "satnr": row[1].value,
            "main_attribute": row[2].value
        }
        models.MainAttribute.objects.create(**data_dict)
    response['msg'] = 'Excel数据写入成功'
    return JsonResponse(response)

# 迭代GP数据
@require_http_methods(["GET"])
def y1_push(request):
    """ 数据推送 """
    response = {}
    try:
        # === 取数 ===
        queryset = models.MainAttribute.objects.filter()
        response['data'] = []
        sql_data = []
        for i in queryset:
            row = response['data'].append(model_to_dict(i))
        for row in response['data']:
            sql_data.append([
                row['id'],
                row['branch_id'],
                row['satnr'],
                row['main_attribute']
            ])
        # === 推送 ===
        conn = gp_connect()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('truncate table ods.ods_ext_strategy_main_product')
        ret = cur.executemany(
            "insert into ods.ods_ext_strategy_main_product(id,branch_id,satnr,main_attribute,create_time,modify_time) "
            "values (%s,%s,%s,%s,current_timestamp,current_timestamp);", sql_data)
        conn.commit()
        conn.close()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

# ======Y2.可比门店表=======
# 按门店ID查询
@require_http_methods(["GET"])
def y2_list(request):
    """ list列表 + 查找数据(模糊查询) """
    response = {}
    try:
        # === 查询开始 ===
        data_dict = {}
        store_id = request.GET.get('store_id')
        if store_id:
            data_dict["store_id__contains"] = store_id
        # === 查询结束 ===
        queryset = models.StoreComparable.objects.filter(**data_dict)
        response['data'] = []
        for i in queryset:
            response['data'].append(model_to_dict(i))
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y2_add(request):
    """ list列表添加与编辑 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        data_dict = {
            "store_id": request.GET.get('store_id'),
            "is_comparable": request.GET.get('is_comparable'),
            "modify_time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        if nid:
            models.StoreComparable.objects.filter(id=nid).update(**data_dict)
        else:
            models.StoreComparable.objects.create(**data_dict)
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y2_delete(request):
    """ list列表删除 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        models.StoreComparable.objects.filter(id=nid).delete()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["POST"])
def y2_upload(request):
    """ Excel上传+表头认证 """
    response = {}
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.打开excel
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.认证表头
    colnames = []
    colnames_real = []
    for col in sheet.iter_cols(max_col=2):
        colnames.append(col[0].value)
    exclude_fields = ('id', 'create_time', 'modify_time')
    params = [f for f in models.StoreComparable._meta.fields if f.name not in exclude_fields]
    for para in params:
        colnames_real.append(para.verbose_name)
    print(colnames, colnames_real)
    if colnames != colnames_real:
        response['msg'] = 'Excel字段名称不匹配'
        return JsonResponse(response)
    # 4.写入数据库
    for row in sheet.iter_rows(min_row=2):
        data_dict = {
            "store_id": row[0].value,
            "is_comparable": row[1].value
        }
        models.StoreComparable.objects.create(**data_dict)
    response['msg'] = 'Excel数据写入成功'
    return JsonResponse(response)

# 迭代GP数据
@require_http_methods(["GET"])
def y2_push(request):
    """ 数据推送 """
    response = {}
    try:
        # === 取数 ===
        queryset = models.StoreComparable.objects.filter()
        response['data'] = []
        sql_data = []
        for i in queryset:
            row = response['data'].append(model_to_dict(i))
        for row in response['data']:
            sql_data.append([
                row['id'],
                row['store_id'],
                row['is_comparable']
            ])
        # === 推送 ===
        conn = gp_connect()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('truncate table ods.ods_ext_store_comparable')
        ret = cur.executemany(
            "insert into ods.ods_ext_store_comparable(id,store_id,is_comparable,create_time,modify_time) "
            "values (%s,%s,%s,current_timestamp,current_timestamp);", sql_data)
        conn.commit()
        conn.close()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

# ======Y3.新老门店映射表=======
# 按新门店ID查询
@require_http_methods(["GET"])
def y3_list(request):
    """ list列表 + 查找数据(模糊查询) """
    response = {}
    try:
        # === 查询开始 ===
        data_dict = {}
        new_store_id = request.GET.get('new_store_id')
        if new_store_id:
            data_dict["new_store_id__contains"] = new_store_id
        # === 查询结束 ===
        queryset = models.StoreMapping.objects.filter(**data_dict)
        response['data'] = []
        for i in queryset:
            response['data'].append(model_to_dict(i))
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y3_add(request):
    """ list列表添加与编辑 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        data_dict = {
            "old_store_id": request.GET.get('old_store_id'),
            "new_store_id": request.GET.get('new_store_id'),
            "new_branch_id": request.GET.get('new_branch_id'),
            "modify_time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        if nid:
            models.StoreMapping.objects.filter(id=nid).update(**data_dict)
        else:
            models.StoreMapping.objects.create(**data_dict)
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["GET"])
def y3_delete(request):
    """ list列表删除 """
    response = {}
    try:
        # === 开始 ===
        nid = request.GET.get('id')
        models.StoreMapping.objects.filter(id=nid).delete()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["POST"])
def y3_upload(request):
    """ Excel上传+表头认证 """
    response = {}
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.打开excel
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.认证表头
    colnames = [] # excel表头
    colnames_real = [] # 数据表表头
    for col in sheet.iter_cols(max_col=3):
        colnames.append(col[0].value)
    exclude_fields = ('id', 'create_time', 'modify_time')
    params = [f for f in models.StoreMapping._meta.fields if f.name not in exclude_fields]
    for para in params:
        colnames_real.append(para.verbose_name)
    print(colnames, colnames_real)
    if colnames != colnames_real:
        response['msg'] = 'Excel字段名称不匹配'
        return JsonResponse(response)
    # 4.写入数据库
    for row in sheet.iter_rows(min_row=2):
        data_dict = {
            "old_store_id": row[0].value,
            "new_store_id": row[1].value,
            "new_branch_id": row[2].value
        }
        models.StoreMapping.objects.create(**data_dict)
    response['msg'] = 'Excel数据写入成功'
    return JsonResponse(response)

# 迭代GP数据
@require_http_methods(["GET"])
def y3_push(request):
    """ 数据推送 """
    response = {}
    try:
        # === 取数 ===
        queryset = models.StoreMapping.objects.filter()
        response['data'] = []
        sql_data = []
        for i in queryset:
            row = response['data'].append(model_to_dict(i))
        for row in response['data']:
            sql_data.append([
                row['id'],
                row['old_store_id'],
                row['new_store_id'],
                row['new_branch_id']
            ])
        # === 推送 ===
        conn = gp_connect()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('truncate table ods.ods_ext_storemapping')
        ret = cur.executemany(
            "insert into ods.ods_ext_storemapping(id,old_store_id,new_store_id,new_branch_id,create_date,modify_time) "
            "values (%s,%s,%s,%s,current_timestamp,current_timestamp);", sql_data)
        conn.commit()
        conn.close()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

# ======Y4.指定加盟店表=======
@require_http_methods(["GET"])
def y4_list(request):
    """ list列表 + 查找数据(模糊查询) """
    response = {}
    try:
        # === 查询开始 ===
        data_dict = {}
        # === 查询结束 ===
        queryset = models.StoreAssign.objects.all()
        response['data'] = []
        for i in queryset:
            response['data'].append(model_to_dict(i))
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)

@require_http_methods(["POST"])
def y4_upload(request):
    """ Excel上传+表头认证 """
    response = {}
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.打开excel
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.认证表头
    colnames = [] # excel表头
    colnames_real = [] # 数据表表头
    for col in sheet.iter_cols(max_col=1):
        colnames.append(col[0].value)
    exclude_fields = ('id', 'create_time', 'modify_time')
    params = [f for f in models.StoreAssign._meta.fields if f.name not in exclude_fields]
    for para in params:
        colnames_real.append(para.verbose_name)
    print(colnames, colnames_real)
    if colnames != colnames_real:
        response['msg'] = 'Excel字段名称不匹配'
        return JsonResponse(response)
    # 4.写入数据库
    # 写之前清空数据表
    models.StoreAssign.objects.all().delete()
    for row in sheet.iter_rows(min_row=2):
        data_dict = {
            "store_id": row[0].value
        }
        models.StoreAssign.objects.create(**data_dict)
    response['msg'] = 'Excel数据写入成功'
    return JsonResponse(response)

# 迭代GP数据
@require_http_methods(["GET"])
def y4_push(request):
    """ 数据推送 """
    response = {}
    try:
        # === 取数 ===
        queryset = models.StoreAssign.objects.filter()
        response['data'] = []
        sql_data = []
        for i in queryset:
            row = response['data'].append(model_to_dict(i))
        for row in response['data']:
            sql_data.append([
                row['id'],
                row['store_id']
            ])
        # === 推送 ===
        conn = gp_connect()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('truncate table ods.ods_ext_storeassign')
        ret = cur.executemany(
            "insert into ods.ods_ext_storeassign(id,store_id,create_date,modify_time) "
            "values (%s,%s,current_timestamp,current_timestamp);", sql_data)
        conn.commit()
        conn.close()
        # === 结束 ===
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = str(e)
    return JsonResponse(response)
